from queue import Empty
import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import fillpdf
from fillpdf import fillpdfs
from openpyxl import load_workbook
from PyPDF2 import PdfFileMerger
import pdfkit
from win32com import client
import xlsxwriter
from io import BytesIO


merger = PdfFileMerger()

def highlight_cols(x):              
    df = x.copy()
    df.loc[:, :] = 'background-color: lightgrey' 
    df.style.hide_index()
    return df 

st.image('zizzl health logo 22.png')

st.title("CSA Roll Up")

workbook = load_workbook(filename = 'ChartC.xlsx')
sheet = workbook.active

#st.write(fillpdfs.get_form_fields('CSA_template.pdf'))

plans = pd.read_csv('plans.csv')
counties = pd.read_csv('counties.csv')
counties_to_plans = pd.read_csv('county to plans.csv')
plans=plans.dropna(subset = 'name')
prices = pd.read_csv('pricings.csv')
df_cols = ['Class', "EE's" ,'People', 'Age 30 Premium', 'County', 'Benchmark Plan', 'AV', 'Single Deductible', 'SBC']
chartC = pd.DataFrame(columns=df_cols)
#st.write(chartC)


join = plans.merge(counties_to_plans, on = 'id', how = 'inner')
#tempjoin = counties_to_plans[['county_id', 'id']]
#prices = prices.merge(tempjoin, on = 'id', how = 'inner')

#prices = plans.merge(prices, on = 'id', how = 'inner')

#join = pd.merge(classed_file_df[['FIPS','Class']], zip_to_fips_df, on = 'FIPS', how = 'inner')

#st.write(plans)
#st.write(counties)
#st.write(counties_to_plans)
#st.write(prices)



company_name = st.text_input('Enter Company Name')

count = st.number_input('Enter Number of Classes')


employee_total = 0
covered_total = 0
zH_monthly_premium_total = 0
current_monthly_premium_total = 0
difference_monthly_premium_total = 0
filtered_in_data = pd.DataFrame()

Current_Monthly_Cost = st.number_input('Current Average Monthly Cost:')

#st.write(counties[['name','state_id']])

class_letter = ''
row_index = 11
col_index = 1


if(count > 0 and Current_Monthly_Cost > 0):
    for i in range(int(count)): 
        st.subheader('Class ')
        class_letter = st.text_input('Class Letter'+str(i+1))
        if class_letter is not None:
            
            Emp = st.number_input('Employees: Class ' + class_letter)
            Covered = st.number_input('Covered: Class '+ class_letter)
            zH_Premium = st.number_input('Premium: Class '+ class_letter)
            state = st.selectbox('Pick State - Class ' + class_letter, [ 'AK', 'AL', 'AR', 'AZ', 'CA', 'CO', 'CT', 'DC', 'DE', 'FL', 'GA','HI', 'IA', 'ID', 'IL', 'IN', 'KS', 'KY', 'LA', 'MA', 'MD', 'ME','MI', 'MN', 'MO', 'MS', 'MT', 'NC', 'ND', 'NE', 'NH', 'NJ', 'NM','NV', 'NY', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX','UT', 'VA', 'VT', 'WA', 'WI', 'WV', 'WY'])
            #filtered_in_data = pd.concat([filtered_in_data, counties[counties['state_id'] == state]], axis=0)
            #st.write(counties[counties['state_id'] == state])
            county = st.selectbox('Pick A County - Class ' + class_letter, counties['name'][counties['state_id'] == state])
            county_id = counties['id'][counties['name'] == county].reset_index(drop = True)
            #st.write(county_id[0])
            temp = join[join['county_id'] == county_id[0]]
            #st.write(temp)
            plan = st.selectbox('Pick a Plan for Class ' + class_letter, temp['name'][temp['off_market'] == True])

            premium_30 = st.number_input('Age 30 Premium for Class '+ class_letter)

            
            st.write('Benchmark Plan: ', plan)
            st.write('County: ', county)
            temp2 = temp[temp['name'] == plan].reset_index()
            st.write(temp2)
            av = temp2['actuarial_value'][0]
            st.write('AV', temp2['actuarial_value'][0])
            single_deductible = temp2['individual_medical_deductible'][0]
            single_deductible = single_deductible[single_deductible.find(':')+1 : single_deductible.find('/')]
            st.write('Single Deductible: ', temp2['individual_medical_deductible'][0])
            temp3 = prices[prices['rating_area_id'].str.slice(0,2,1) == state]
            #st.write(temp3)
            sbc = temp2['summary_of_benefits_url'][0] 
            st.write('SBC: ', temp2['summary_of_benefits_url'][0])
            #st.write('Age 30 Premium: ', temp4[['age_30','id', 'rating_area_id']][temp4['rating_area_id'].str.slice(0,2,1) == state])
            chartC = chartC.append({'Class' : class_letter , "EE's" : round(Emp), 'People' : round(Covered), 'Age 30 Premium': round(premium_30, 2), 'County': county, 'Benchmark Plan' : plan, 'AV': round(av, 2), 'Single Deductible' : single_deductible, 'SBC': sbc},  ignore_index = True)
            sheet.cell(row_index, col_index, value = class_letter)
            sheet.cell(row_index, col_index+1, value = Emp)
            sheet.cell(row_index, col_index+2, value = Covered)
            sheet.cell(row_index, col_index+3, value = round(premium_30, 2))
            sheet.cell(row_index, col_index+4, value = county)
            sheet.cell(row_index, col_index+5, value = '=HYPERLINK("{}", "{}")'.format(sbc, plan))
            sheet.cell(row_index, col_index+6, value = av)
            sheet.cell(row_index, col_index+7, value = single_deductible)
            sheet.cell(row_index, col_index+8, value = '')
            row_index+=1
            #filtered_out_data['carrier_name'][filtered_out_data['id'].str.slice(5, 7, 1) == i]]
        

        current_temp_premium = Current_Monthly_Cost * Emp

        difference_temp_premium = zH_Premium - current_temp_premium

        current_monthly_premium_total += current_temp_premium

        difference_monthly_premium_total += difference_temp_premium
        

        zH_monthly_premium_total += zH_Premium
        employee_total += Emp
        covered_total += Covered
    st.table(chartC.style.apply(highlight_cols, axis = None))    
    
    zH_yearly_premium_total = zH_monthly_premium_total * 12
    zh_average_monthly_cost = zH_monthly_premium_total / employee_total
    
    current_yearly_premium_total = current_monthly_premium_total * 12

    difference_yearly_premium_total = zH_yearly_premium_total - current_yearly_premium_total
    difference_average_monthly_cost = zh_average_monthly_cost - Current_Monthly_Cost

    st.subheader('Total: ')
    st.write('Employee Total: ', employee_total)
    st.write('Covered Total: ', covered_total)
    st.write('zH Monthly Premium Total: ', zH_monthly_premium_total)
    st.write('zH Yearly Premium Total: ', zH_yearly_premium_total)
    st.write('zH Average Monthly Cost: ',zh_average_monthly_cost)

    st.write('Current Monthly Premium Total' ,current_monthly_premium_total)
    st.write('Current Yearly Premium Total', current_yearly_premium_total)
    st.write('Difference Monthly Premium Total: ', difference_monthly_premium_total)
    st.write('Difference Yearly Premium Total: ', difference_yearly_premium_total)
    st.write('Difference Average Monthly Cost: ' ,difference_average_monthly_cost)

    st.write('% Difference: ', (difference_average_monthly_cost/Current_Monthly_Cost)*100, '%')


    Admin_Fee = st.number_input('Annual Admin Fee:')
    Consulting_Fee = st.number_input('Annual Consulting Fee:')

    Annual_zH_Cost = Admin_Fee + Consulting_Fee + zH_yearly_premium_total

    st.write('Annual zH Cost: ', Annual_zH_Cost)

    zH_Cost_Difference =  ((Annual_zH_Cost - current_yearly_premium_total)/current_yearly_premium_total)*100

    st.write('zH Cost Difference: ', zH_Cost_Difference, '%')

    data_dict = {
    "Company Name":company_name,
    'Covered EEs': employee_total,
    'Total Covered Lives': covered_total,
    "Text1":zH_monthly_premium_total,
    "Text2":zH_yearly_premium_total,
    "Text3":zh_average_monthly_cost,
    "Text4":current_monthly_premium_total,
    "Text5": current_yearly_premium_total,
    "Text6":Current_Monthly_Cost,
    "Text7":difference_monthly_premium_total,
    "Text8":difference_yearly_premium_total,
    "Text9":difference_average_monthly_cost,
    "Text10":zH_Cost_Difference,
    "Text11":zH_Cost_Difference,
    "Text12":zH_Cost_Difference,
    "$":Admin_Fee,
    "$$":Consulting_Fee,
    "$$$":Annual_zH_Cost,
    "%":str(zH_Cost_Difference) + '%'
    }


    workbook.save('test.xlsx')

    test = pd.read_excel('test.xlsx')

    fillpdfs.write_fillable_pdf('CSA_template.pdf', 'new.pdf', data_dict)
    fig, ax =plt.subplots(figsize=(12,4))
    ax.axis('tight')
    ax.axis('off')
    the_table = ax.table(cellText=chartC.values,colLabels=chartC.columns,loc='center')


    with open("new.pdf", "rb") as pdf_file:
        PDFbyte = pdf_file.read()

    st.download_button(label="Download CSA",
                        data=PDFbyte,
                        file_name="RollUpCSA.pdf",
                        mime='application/octet-stream') 

    with open('test.xlsx', 'rb') as my_file:
        st.download_button(label = 'Download Chart C', data = my_file, file_name = 'Chart_C.xlsx', mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')